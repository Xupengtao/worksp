# -*- coding: utf-8 -*-
"""
Created on Tue Oct 16 20:07:24 2018

@author: wolai
"""
from openpyxl import load_workbook
from openpyxl import Workbook

class ExcelPro:
    def __init__(self,filepath = ''): 
        self.filepath = filepath
        if filepath != '':
            self.wb = load_workbook(filepath)
        else:
            self.wb = Workbook()
            self.wb.active
        self.ws = []
        for sheetname in self.wb.sheetnames:
            self.ws.append(self.wb[sheetname])
        self.setsheet(0)
            
    def creatsheet(self,pos = 0,name = ''):
        if name == '':
            self.ws.append(self.wb.create_sheet())
        else:
            if pos <= len(self.ws):
                self.ws.insert(pos,self.wb.create_sheet(name,pos))
            else:
                self.ws.append(self.wb.create_sheet(name,pos))
                
    def changesheet(self,pos,name):
        if pos <= len(self.ws):
            del self.wb[str(self.ws[pos])[12:-2]]
            self.ws[pos] = self.wb.create_sheet(name,pos)
            
    def deletesheet(self,*,pos = 0,name = ''):
        if pos != 0 and pos < len(self.ws):
            del self.wb[str(self.ws[pos])[12:-2]]
            del self.ws[pos]
            return 0
        if name != '':
            del self.wb[name]
            i = 0
            while i < len(self.ws):
                if str(self.ws[i])[12:-2] == name:
                    del self.ws[i]
                    i = i-1
                i = i+1
    
    def changetabcolor(self,*,pos = 0,color = '888888'):
        if pos < len(self.ws): 
            self.ws[pos].sheet_properties.tabColor = color
    
    def setsheet(self,pos):
        if pos < len(self.ws): 
            self.sheetpos = pos
            self.max_row = self.ws[self.sheetpos].max_row
            self.max_column = self.ws[self.sheetpos].max_column
            return 0
        else:
            return -1
    
    def readrow(self, row):
        rowdata = [0 for x in range(self.max_column)]
        if row <= self.max_row:
            for col in range(self.max_column):
                rowdata[col] = self.readcell(row,col+1)
            return rowdata
        else:
            return -1
            
    def readcolumn(self, column):
        rowdata = [0 for x in range(self.max_row)]
        if column <= self.max_column:
            for row in range(self.max_row):
                rowdata[row] = self.readcell(row+1,column)
            return rowdata
        else:
            return -1
        
    def writerow(self, row = 0, *, rowlist = [], rowdict = {}):
        try:
            if row == 0 and rowdict != {}:
                self.ws[self.sheetpos].append(rowdict)
            elif row != 0 and rowlist != []:
                for col in range(len(rowlist)):
                    self.writecell(row,col+1,rowlist[col])
        except:
            return -1
    
    def readcell(self, row, column):
        try:
            return self.ws[self.sheetpos].cell(row = row, column = column).value
        except:
            return -1
    
    
    def writecell(self, row, column, value):
        try:
            self.ws[self.sheetpos].cell(row = row,column = column).value = value
#           sheet.cell(row=3,column=5,value=100)
        except:
            return -1
    
    def savefile(self,filepath = ''):
        try:
            if filepath == '' and self.filepath != '':
                self.wb.save(self.filepath)
            elif filepath != '':
                self.wb.save(filepath)
        except:
            return -1